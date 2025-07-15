from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet

from django.shortcuts import render
from django.views.generic import TemplateView
from django.http import HttpResponse
from django.urls import get_resolver
from django.utils import timezone
from django.db.models import Sum,F
from django.db.models.functions import ExtractMonth, ExtractQuarter
from django.db import transaction
from django.db import connection

import openpyxl
from datetime import datetime

from .models import (
    Account,
    Invoice,
    Rent,
    Room,
    RoomType,
    Staff,
    Customer,
    Service,
    ServiceUsage
)

from .serializers import (    
    LoginSerializer,
    RentSerializer,
    RoomSerializer,
    InvoiceSerializer,
    AccountSerializer,
    RoomTypeSerializer,
    StaffSerializer,
    CustomerSerializer,
    ServiceUsageSerializer,
    ServiceSerializer,
)

# Create your views here.
# Add api here
class AccountCreateAPIView(APIView):
    def post(self, request, *args, **kwargs):
        with transaction.atomic():
            serializer = AccountSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response({'message': 'Account created successfully'}, status=status.HTTP_201_CREATED)
        return Response({'message': 'Account creation failed', 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
    
class LoginView(APIView):
    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.validated_data['user']
            user_data = AccountSerializer(user).data  
            return Response({
                "message": "Đăng nhập thành công",
                **user_data,
                'role': user.ma_nhan_vien.role
            }, status=status.HTTP_200_OK)
        return Response({"message": "Đăng nhập thất bại", "errors": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
    
class BookingViewSet(ModelViewSet):
    queryset = Rent.objects.all().order_by('ma_thue')
    serializer_class = RentSerializer

    def create(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                response = super().create(request, *args, **kwargs)
                return Response({
                    "message": "Đặt phòng thành công",
                    "data": response.data
                }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({
                "message": "Đặt phòng thất bại",
                "error": str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['put'], url_path='extend')
    def extend_booking(self, request, pk=None):
        with transaction.atomic():
            rent = self.get_object()
            new_date = request.data.get('ngay_tra', None)

            if not new_date:
                return Response({"error": "Vui lòng cung cấp ngay_tra."}, status=400)
            new_date = datetime.strptime(new_date, '%Y-%m-%d').date()

            if new_date <rent.ngay_nhan:
                return Response({"error": "Ngày trả không hợp lệ."}, status=400)
            
            if new_date == rent.ngay_tra:
                return Response({"error": "Ngày trả không thay đổi."}, status=400)
            
            # Kiểm tra phòng đã có người đặt/ở sau thời gian mới chưa?
            conflict = Rent.objects.filter(
                ma_phong=rent.ma_phong,
                ngay_nhan__lte=new_date,
                trang_thai__in=['Đã đặt', 'Đang ở']
            ).exclude(pk=rent.pk).exists()
            if conflict:
                return Response({"error": "Phòng đã có người đặt sau thời gian này."}, status=400)
            rent.ngay_tra = new_date
            try:
                rent.save()
            except Exception as e:
                return Response({'message': 'Gia hạn thất bại', "error": str(e)}, status=400)
            return Response({"message": "Gia hạn thành công."})

    def destroy(self, request, *args, **kwargs):
        if self.get_object().trang_thai == 'Đang ở' or self.get_object().trang_thai == 'Đã trả':
            return Response({"message": "Không thể xóa đặt phòng đang ở hoặc đã trả"}, status=status.HTTP_400_BAD_REQUEST)
        
        with transaction.atomic():    
            super().destroy(request, *args, **kwargs)
            return Response({
                "message": "Xóa đặt phòng thành công"
            }, status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['put'], url_path='check-in')
    def check_in(self, request, pk=None):
        with transaction.atomic():
            rent = self.get_object()
            rent.trang_thai = 'Đang ở'
            room = rent.ma_phong  # Đây là instance của model Room
            room.trang_thai = 'Đang sử dụng'
            room.save()  # ← Ghi xuống database đúng cách
            rent.save()
            return Response({"message": "Check-in thành công."})
    
    @action(detail=True, methods=['put'], url_path='update-booking')
    def update_booking(self, request, pk=None):
        with transaction.atomic():
            rent = self.get_object()
            if rent.trang_thai == 'Đang ở' or rent.trang_thai == 'Đã trả':
                return Response({"error": "Không thể cập nhật đặt phòng đã ở hoặc đã trả."},
                                status=status.HTTP_400_BAD_REQUEST)
            
            serializer = self.get_serializer(rent, data=request.data, partial=True)
            if not serializer.is_valid(raise_exception=True):
                return Response({"error": "Dữ liệu không hợp lệ."}, status=status.HTTP_400_BAD_REQUEST)

            phong = serializer.validated_data.get('ma_phong', None)
            ngay_nhan_new = serializer.validated_data.get('ngay_nhan', None)
            ngay_tra_new = serializer.validated_data.get('ngay_tra', None)
            
            if ngay_tra_new < ngay_nhan_new:
                return Response({"error": "Ngày trả phải sau hoặc bằng ngày nhận."},
                                status=status.HTTP_400_BAD_REQUEST)
            if not ngay_nhan_new or not ngay_tra_new:
                return Response({"error": "Ngày nhận, ngày thuê và ngày trả là bắt buộc."},
                                status=status.HTTP_400_BAD_REQUEST)
            if phong:
                # Check phòng còn trống & không bị trùng
                if phong.trang_thai == 'Trống':
                    conflict=Rent.objects.filter(
                    Q(ma_phong=phong) & 
                    (Q(trang_thai='Đã đặt') | Q(trang_thai='Đang ở')) &
                    ((Q(ngay_tra__gte=ngay_nhan_new) & Q(ngay_nhan__lte=ngay_nhan_new)) |
                    (Q(ngay_tra__gte=ngay_tra_new) & Q(ngay_nhan__lte=ngay_tra_new)) |
                    (Q(ngay_tra__lt=ngay_tra_new) & Q(ngay_nhan__gt=ngay_nhan_new))
                    )).exclude(pk=rent.pk).exists()
                    if conflict:
                        return Response({"error": "Phòng đã có người đặt trong khoảng thời gian này."},
                                        status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({"error": "Phòng hiện không trống."},
                                    status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({"error": "Mã Phòng là bắt buộc."},
                                status=status.HTTP_400_BAD_REQUEST)
            try:
                serializer.save()
                return Response({"message": "Cập nhật đặt phòng thành công", "data": serializer.data})
            except Exception as e:
                return Response({"message": "Cập nhật đặt phòng thất bại", "errors": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
            
    @action(detail=False, methods=['delete'], url_path='clean-up')
    def cleanup_expired_bookings(self, request):
        with transaction.atomic():
            today = timezone.now().date()
            expired = Rent.objects.filter(
                trang_thai='Đã đặt',
                ngay_nhan__lt=today
            )
            count = expired.count()
            try:
                expired.delete()
                return Response({"message": f"Đã xoá {count} đặt phòng quá hạn check-in."})
            except Exception as e:
                return Response({"message": "Xóa đặt phòng thất bại", "errors": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
class InvoiceViewSet(ModelViewSet):
    queryset = Invoice.objects.all().order_by('ma_hoa_don')
    serializer_class = InvoiceSerializer

    def create(self, request, *args, **kwargs):
        with transaction.atomic():
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid()
            invoice = serializer.save()  # Gọi vào `InvoiceSerializer.create()`
            return Response({"message": "Tạo hóa đơn thành công", "ma_hoa_don": invoice.ma_hoa_don}, status=201)
    
    def destroy(self, request, *args, **kwargs):
        with transaction.atomic():
            invoice=self.get_object()
            if invoice.trang_thai == 'Đã thanh toán':
                return Response({"message": "Không thể xóa hóa đơn đã thanh toán"}, status=status.HTTP_400_BAD_REQUEST)
            
            super().destroy(request, *args, **kwargs)
            return Response({"message": "Xóa hóa đơn thành công"}, status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=True, methods=['put'], url_path='payment')
    def payment(self, request, pk=None):
        with transaction.atomic():
            try:
                invoice = self.get_object()
                if invoice.trang_thai != 'Chưa thanh toán':
                    return Response({'message': 'Hoá đơn đã thanh toán'}, status=status.HTTP_400_BAD_REQUEST)
                
                rent_list = invoice.rents.all()
                for rent in rent_list:
                    rent.trang_thai = 'Đã trả'
                    rent.ngay_tra = timezone.now().date()
                    ma_phong=rent.ma_phong
                    ma_phong.trang_thai = 'Trống'
                    ma_phong.save()
                    rent.save()
                
                usage_list = invoice.service_usages.all()
                for usage in usage_list:
                    usage.trang_thai = 'Đã thanh toán'
                    usage.save()

                invoice.trang_thai = 'Đã thanh toán'
                invoice.save()
                return Response({"message": "Thanh toán hóa đơn thành công."})
            except Exception as e:
                return Response({"message": "Thanh toán hóa đơn thất bại", "errors": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class RevenueStatsAPIView(APIView):
    def get(self, request):
        period = request.GET.get('period', 'month')
        year = request.GET.get('year', datetime.now().year)

        invoices = Invoice.objects.filter(ngay_lap_hoa_don__year=year)

        if period == 'month':
            data = invoices.annotate(month=ExtractMonth('ngay_lap_hoa_don')) \
                .values('month') \
                .annotate(doanh_thu=Sum('tong_tien')) \
                .order_by('month')
        elif period == 'quarter':
            data = invoices.annotate(quarter=ExtractQuarter('ngay_lap_hoa_don')) \
                .values('quarter') \
                .annotate(doanh_thu=Sum('tong_tien')) \
                .order_by('quarter')
        else:
            return Response({"error": "Tham số period không hợp lệ."}, status=400)

        return Response({
            "period": period,
            "year": int(year),
            "data": list(data)
        })

class ServiceUsageStatsAPIView(APIView):
    def get(self, request):
        # Lấy tham số từ request
        period = request.GET.get('period', 'month')
        year = request.GET.get('year', datetime.now().year)

        # Lọc các dịch vụ đã thanh toán trong năm
        usages = ServiceUsage.objects.filter(ngay_su_dung__year=year, status='Đã thanh toán')

        # Tính giá trị dịch vụ trước
        usages = usages.annotate(
            total_value=F('so_luong') * F('ma_dich_vu__gia_dich_vu')  # Tính giá trị dịch vụ
        )

        # Phân nhóm theo quý và dịch vụ
        if period == 'quarter':
            data = usages.annotate(quarter=ExtractQuarter('ngay_su_dung')) \
                .values('quarter', 'ma_dich_vu__ten_dich_vu', 'ma_dich_vu__ma_dich_vu') \
                .annotate(
                    total_usage=Sum('so_luong'),
                    total_revenue=Sum('total_value')
                ) \
                .order_by('quarter', 'ma_dich_vu__ma_dich_vu')

        else:
            return Response({"error": "Tham số period không hợp lệ."}, status=400)

        # Tạo cấu trúc kết quả cho dễ sử dụng
        result = {}
        for entry in data:
            quarter = entry['quarter']
            service_name = entry['ma_dich_vu__ten_dich_vu']
            total_usage = entry['total_usage']
            total_revenue = entry['total_revenue']

            # Nếu chưa có quý này trong kết quả, tạo mới
            if quarter not in result:
                result[quarter] = {
                    "total_revenue": 0,  # Khởi tạo doanh thu tổng của quý
                }

            # Thêm doanh thu của dịch vụ vào quý
            result[quarter][service_name] = total_revenue
            result[quarter]["total_revenue"] += total_revenue  # Cộng dồn doanh thu tổng

        # Định dạng kết quả trả về
        final_result = []
        for quarter, services in result.items():
            entry = {'quarter': quarter}
            # Duyệt qua từng dịch vụ trong quý để thêm vào kết quả
            for service_name, revenue in services.items():
                if service_name != "total_revenue":
                    entry[service_name] = revenue
            entry["total_revenue"] = services["total_revenue"]
            final_result.append(entry)

        return Response({
            "period": period,
            "year": int(year),
            "data": final_result
        })

class ExportStatsExcelAPIView(APIView):
    # def get(self, request):
    #     year = request.GET.get('year', datetime.now().year)

    #     # Lấy doanh thu theo tháng
    #     doanh_thu = Invoice.objects.filter(ngay_lap_hoa_don__year=year) \
    #         .annotate(month=ExtractMonth('ngay_lap_hoa_don')) \
    #         .values('month') \
    #         .annotate(doanh_thu=Sum('tong_tien')) \
    #         .order_by('month')

    #     # Lấy sử dụng dịch vụ theo tháng
    #     dich_vu = ServiceUsage.objects.filter(ngay_su_dung__year=year) \
    #         .annotate(month=ExtractMonth('ngay_su_dung')) \
    #         .values('month') \
    #         .annotate(so_luong=Sum('so_luong')) \
    #         .order_by('month')

    #     # Tạo file Excel
    #     wb = openpyxl.Workbook()
    #     ws1 = wb.active
    #     ws1.title = "Doanh thu"

    #     ws1.append(["Tháng", "Doanh thu"])
    #     for row in doanh_thu:
    #         ws1.append([row['month'], row['doanh_thu']])

    #     ws2 = wb.create_sheet(title="Sử dụng dịch vụ")
    #     ws2.append(["Tháng", "Số lượng"])
    #     for row in dich_vu:
    #         ws2.append([row['month'], row['so_luong']])

    #     # Trả file
    #     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #     filename = f"thong_ke_{year}.xlsx"
    #     response['Content-Disposition'] = f'attachment; filename={filename}'
    #     wb.save(response)
    #     return response
    def get(self, request):
        year = int(request.GET.get('year', datetime.now().year))

        # Sheet 1: Hóa đơn
        invoices = Invoice.objects.filter(ngay_lap_hoa_don__year=year).select_related("ma_nhan_vien")

        # Sheet 2: Khách hàng
        customers = Customer.objects.all()

        # Sheet 3: Thuê phòng
        rents = Rent.objects.filter(ngay_thue__year=year).select_related("ma_khach_hang", "ma_nhan_vien", "ma_phong")

        # Sheet 4: Sử dụng dịch vụ
        services = ServiceUsage.objects.filter(ngay_su_dung__year=year).select_related("ma_khach_hang", "ma_dich_vu")

        wb = openpyxl.Workbook()

        # Hóa đơn
        ws1 = wb.active
        ws1.title = "Hoa don"
        ws1.append(["Mã hóa đơn", "Ngày lập", "Trạng thái", "Tổng tiền", "Người lập"])
        for i in invoices:
            ws1.append([i.ma_hoa_don, i.ngay_lap_hoa_don, i.trang_thai, float(i.tong_tien), i.ma_nhan_vien.ho_ten if i.ma_nhan_vien else ''])

        # Khách hàng
        ws2 = wb.create_sheet("Khach hang")
        ws2.append(["Mã KH", "Tên", "Địa chỉ", "SĐT"])
        for c in customers:
            ws2.append([c.ma_khach_hang, c.ten_khach_hang, c.dia_chi, c.so_dien_thoai])

        # Thuê phòng
        ws3 = wb.create_sheet("Thue phong")
        ws3.append(["Mã thuê", "Khách", "Phòng", "Ngày thuê", "Ngày nhận", "Ngày trả", "Trạng thái", "Nhân viên"])
        for r in rents:
            ws3.append([
                r.ma_thue,
                r.ma_khach_hang.ten_khach_hang if r.ma_khach_hang else '',
                r.ma_phong.ma_phong if r.ma_phong else '',
                r.ngay_thue,
                r.ngay_nhan,
                r.ngay_tra,
                r.trang_thai,
                r.ma_nhan_vien.ho_ten if r.ma_nhan_vien else ''
            ])

        # Dịch vụ sử dụng
        ws4 = wb.create_sheet("Dich vu")
        ws4.append(["Mã SDDV", "Khách hàng", "Dịch vụ", "Số lượng", "Ngày SD", "Giá", "Trạng thái"])
        for s in services:
            ws4.append([
                s.ma_sddv,
                s.ma_khach_hang.ten_khach_hang if s.ma_khach_hang else '',
                s.ma_dich_vu.ten_dich_vu if s.ma_dich_vu else '',
                s.so_luong,
                s.ngay_su_dung,
                float(s.ma_dich_vu.gia_dich_vu) if s.ma_dich_vu else '',
                s.status
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=bao_cao_{year}.xlsx'
        wb.save(response)
        return response
    
class AccountViewSet(ModelViewSet):
    queryset = Account.objects.all().order_by('id')
    serializer_class = AccountSerializer

    def create(self, request, *args, **kwargs):
        with transaction.atomic():
            response = super().create(request, *args, **kwargs)
            return Response({
                "message": "Tạo tài khoản thành công",
                "data": response.data
            }, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        with transaction.atomic():
            response = super().update(request, *args, **kwargs)
            return Response({
                "message": "Cập nhật tài khoản thành công",
                "data": response.data
            }, status=status.HTTP_200_OK)
        
    def destroy(self, request, *args, **kwargs):
        with transaction.atomic():
            super().destroy(request, *args, **kwargs)
            return Response({
                "message": "Xóa tài khoản thành công"
            }, status=status.HTTP_204_NO_CONTENT)  

class RoomTypeViewSet(ModelViewSet):
    queryset = RoomType.objects.all().order_by('ma_loai')
    serializer_class = RoomTypeSerializer

    def create(self, request, *args, **kwargs):
        with transaction.atomic():
            response = super().create(request, *args, **kwargs)
            return Response({
                "message": "Tạo loại phòng thành công",
                "data": response.data
            }, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        with transaction.atomic():
            response = super().update(request, *args, **kwargs)
            return Response({
                "message": "Cập nhật loại phòng thành công",
                "data": response.data
            }, status=status.HTTP_200_OK)
    
    def destroy(self, request, *args, **kwargs):
        with transaction.atomic():
            super().destroy(request, *args, **kwargs)
            return Response({
                "message": "Xóa loại phòng thành công"
            }, status=status.HTTP_204_NO_CONTENT)

class RoomViewSet(ModelViewSet):
    queryset = Room.objects.all().order_by('ma_phong')
    serializer_class = RoomSerializer

    def create(self, request, *args, **kwargs):
        with transaction.atomic():
            response = super().create(request, *args, **kwargs)
            return Response({
                "message": "Tạo phòng thành công",
                "data": response.data
            }, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        with transaction.atomic():
            response = super().update(request, *args, **kwargs)
            return Response({
                "message": "Cập nhật phòng thành công",
                "data": response.data
            }, status=status.HTTP_200_OK)
    
    def destroy(self, request, *args, **kwargs):
        with transaction.atomic():
            super().destroy(request, *args, **kwargs)
            return Response({
                "message": "Xóa phòng thành công"
            }, status=status.HTTP_204_NO_CONTENT)

class StaffViewSet(ModelViewSet):
    queryset = Staff.objects.all().order_by('ma_nhan_vien')
    serializer_class = StaffSerializer

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        return Response({
            "message": "Tạo nhân viên thành công",
            "data": response.data
        }, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        return Response({
            "message": "Cập nhật nhân viên thành công",
            "data": response.data
        }, status=status.HTTP_200_OK)
    
    def destroy(self, request, *args, **kwargs):
        instance=self.get_object()
        try:
            account = instance.account
            if account:
                account.delete()
        except Exception:
            pass  # Không có tài khoản nào liên kết, không cần xoá

        super().destroy(request, *args, **kwargs)
        return Response({
            "message": "Xóa nhân viên thành công"
        }, status=status.HTTP_204_NO_CONTENT)

class CustomerViewSet(ModelViewSet):
    queryset = Customer.objects.all().order_by('ma_khach_hang')
    serializer_class = CustomerSerializer

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        return Response({
            "message": "Tạo khách hàng thành công",
            "data": response.data
        }, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        return Response({
            "message": "Cập nhật khách hàng thành công",
            "data": response.data
        }, status=status.HTTP_200_OK)
    
    def destroy(self, request, *args, **kwargs):
        super().destroy(request, *args, **kwargs)
        return Response({
            "message": "Xóa khách hàng thành công"
        }, status=status.HTTP_204_NO_CONTENT)
    
class ServiceUsageViewSet(ModelViewSet):
    queryset = ServiceUsage.objects.all().order_by('ma_sddv')
    serializer_class = ServiceUsageSerializer

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        return Response({
            "message": "Tạo sử dụng dịch vụ thành công",
            "data": response.data
        }, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        return Response({
            "message": "Cập nhật sử dụng dịch vụ thành công",
            "data": response.data
        }, status=status.HTTP_200_OK)
    
    def destroy(self, request, *args, **kwargs):
        super().destroy(request, *args, **kwargs)
        return Response({
            "message": "Xóa sử dụng dịch vụ thành công"
        }, status=status.HTTP_204_NO_CONTENT)

class ServiceViewSet(ModelViewSet):
    queryset = Service.objects.all().order_by('ma_dich_vu')
    serializer_class = ServiceSerializer

    def create(self, request, *args, **kwargs):
        with transaction.atomic():
            response = super().create(request, *args, **kwargs)
            return Response({
                "message": "Tạo dịch vụ thành công",
                "data": response.data
            }, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        with transaction.atomic():
            response = super().update(request, *args, **kwargs)
            return Response({
                "message": "Cập nhật dịch vụ thành công",
                "data": response.data
            }, status=status.HTTP_200_OK)
    
    def destroy(self, request, *args, **kwargs):
        with transaction.atomic():
            super().destroy(request, *args, **kwargs)
            return Response({
                "message": "Xóa dịch vụ thành công"
            }, status=status.HTTP_204_NO_CONTENT)
#Base index view
class IndexView(TemplateView):
    template_name = "index.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Check database connection
        try:
            connection.ensure_connection()
            context['db_status'] = "✅ Connected"
        except Exception as e:
            context['db_status'] = f"❌ Error: {e}"

        # List all registered URL patterns
        resolver = get_resolver()
        all_urls = []
        for pattern in resolver.url_patterns:
            if hasattr(pattern, 'pattern'):
                path = str(pattern.pattern)
                all_urls.append(path)
        context['all_urls'] = all_urls

        return context